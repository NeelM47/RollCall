import openpyxl
import json
from datetime import datetime, timedelta
import pathlib
from icecream import ic

ic.configureOutput(prefix=f'Debug | ', includeContext=True)

year = 2024
month = 11  # November

# Load data from JSON files
with open("public/names.json", "r") as file:
    names_data = json.load(file)

with open("data/data.json", "r") as file:
    attendance_data = json.load(file)

# Extract names and attendance data
names_set = {person['name'] for person in names_data}  # Set for fast lookup
names_with_dates = {}

# Create a dictionary where names are keys and days are values
for person in attendance_data:
    name = person['name']
    day = datetime.fromisoformat(person['day'].replace("Z", "+00:00")).day  # Convert day to just the day of the month
    names_with_dates.setdefault(name, []).append(day)

def get_days(year, month):
    # Start with the first day of the month
    start_date = datetime(year, month, 1)

    # Find the first Tuesday of the month
    first_tuesday = start_date + timedelta(days=(1 - start_date.weekday()) % 7)

    first_thursday = start_date + timedelta(days=(3 - start_date.weekday()) % 7)

    days = []

    # Loop through the month and find all Tuesdays
    current_day_tuesday = first_tuesday
    current_day_thursday = first_thursday
    while current_day_tuesday.month == month or current_day_thursday.month == month:
        if current_day_tuesday.month == month:
            days.append(current_day_tuesday.strftime("%d"))
            current_day_tuesday += timedelta(weeks=1)

        if current_day_thursday.month == month:
            days.append(current_day_thursday.strftime("%d"))
            current_day_thursday += timedelta(weeks=1)

        #days.append(current_day.strftime("%d"))
        #days.append(current_day.strftime("%Y-%m-%d"))
        #current_day += timedelta(weeks=1)  # Move to the next Tuesday

    days.sort(key=lambda x: int(x))

    return days

days_in_month = get_days(year, month)

# Define the directory and check for existing xlsx files
directory = pathlib.Path("data")
xlsx_files = list(directory.glob('*.xlsx'))

# Check if there's an existing file; create one if not
if not xlsx_files:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Add header row (names and days of the month)
    ws.cell(row=1, column=1, value="Names")

    #for i, day in enumerate(days_in_month, start=1):
    #    ws.cell(row=1, column=i + 1, value=day)

    for i in range(1, 32):  # Up to 31 days, you can modify it dynamically if needed
        ws.cell(row=1, column=i + 1, value=str(i))

    # Fill in names in the first column
    for i, name in enumerate(names_set, start=2):
        ws.cell(row=i, column=1, value=name)

    wb.save(f"data/attendance_sheet_{month}.xlsx")

else:
    wb = openpyxl.load_workbook(f"data/attendance_sheet_{month}.xlsx")
    ws = wb.active

    # Iterate through rows and cells to mark attendance
    for row in ws.iter_rows(min_row=2, min_col=1):
        name_cell = row[0]
        if name_cell.value in names_with_dates:  # Only process if the name exists in the attendance data
            for cell in row[1:]:  # Iterate through days (columns)
                current_day = cell.column - 1
                # Check if the current_day matches any day this person attended
                for day_attended in names_with_dates[name_cell.value]:
                    if current_day == day_attended:
                        cell.value = "✓"  # Mark attendance

    wb.save(f"data/attendance_sheet_{month}.xlsx")

